# -*- coding: utf-8 -*-
"""
ocr_bom_pipeline.py - Groupe Bayoudh Metal

Pipeline OCR dédié au formulaire Nomenclature (BOM) ERPNext.
Suit la même architecture que ocr_article_pipeline.py.

Endpoints exposés :
  - pipeline_bom(file_url, source_doctype)  → async, retourne job_id
  - get_ocr_bom_statut(job_id)              → polling résultat

Job interne :
  - _executer_pipeline_bom_job(...)

Champs ERPNext remplis (header) :
  item, item_name, company, quantity, uom, currency,
  rm_cost_as_per, routing, transfer_material_against,
  conversion_rate, is_active, is_default, scrap_percentage

Composants (child table items) :
  item_code, item_name, description, qty, uom,
  qty_per_unit, stock_uom, conversion_factor, rate

──────────────────────────────────────────────────────────────────────
CORRECTIF (v9.4) :
  - Unification de la détection "document rejeté" (chèque / traite /
    fiche article / facture / BL / ...) dans _detecter_rejet_categorie().
  - Cette détection est désormais retentée sur le texte issu de l'OCR
    renforcé (_ocr_enhanced) quand la 1ère passe ne trouve ni signal BOM
    ni pattern de rejet. Avant ce correctif, l'OCR renforcé ne servait
    qu'à chercher des signaux BOM positifs : un chèque/traite mal
    scanné (texte 1ère passe trop pauvre) tombait alors, à tort, sur le
    message générique "document inconnu" au lieu du message précis
    "Chèque" / "Traite".
  - L'appel à extraire_champs_bom() est sécurisé : en cas d'exception,
    on retente une classification via _detecter_rejet_categorie() avant
    d'abandonner sur une erreur générique (corrige le crash observé sur
    une "Fiche Article.pdf" qui remontait "Erreur interne du pipeline").
──────────────────────────────────────────────────────────────────────
"""

import frappe
import os
import re
import json
import uuid
from werkzeug.utils import secure_filename
import re as _re

# ──────────────────────────────────────────────────────────────────────
# SIGNAUX DE DÉTECTION
# ──────────────────────────────────────────────────────────────────────

_SIGNAUX_BOM = [
    "nomenclature", "bom", "bill of material",
    "composants", "composant",
    "quantité produite", "quantite produite",
    "matière première", "matiere premiere",
    "taux de perte", "coût matière",
    "coût total", "cout total",
    "résumé des coûts", "resume des couts",
    "paramètres de fabrication",
]
_SIGNAUX_BOM_FORTS = [
    "nomenclature", "bom", "bill of material",
    "quantité produite", "quantite produite",
    "composants",
]
_MOTS_ARTICLE_STRUCTURE = [
    "fiche article", "fiche technique", "datasheet", "fiche produit",
    "groupe d'article", "unité de mesure par défaut", "unité stock",
    "référence article", "prix unitaire ht", "prix de vente ht",
    "prix achat ht", "part no", "part number", "sku",
]

# ── Patterns de rejet : identifient sans ambiguïté un document NON-BOM ──
_PATTERNS_REJET_IMMEDIAT_BOM = [p for p in [
    # Traite / Lettre de change
    _re.compile(r'\btireur\b', _re.IGNORECASE),                         # 0
    _re.compile(r'\bdomiciliation\b', _re.IGNORECASE),                  # 1
    _re.compile(r'\blettre\s*de\s*change\b', _re.IGNORECASE),           # 2
    _re.compile(r'\bvaleur\s*re[çc]ue?\b', _re.IGNORECASE),             # 3
    _re.compile(r"\bà\s*l[''`]?\s*ordre\s*de\b", _re.IGNORECASE),       # 4
    _re.compile(r'\béch[eé]ance\b', _re.IGNORECASE),                    # 5
    _re.compile(r'\btir[eé]\b', _re.IGNORECASE),                        # 6
    _re.compile(r'\baval\b', _re.IGNORECASE),                           # 7
    _re.compile(r'\btraite\b', _re.IGNORECASE),                         # 8
    _re.compile(r'\bacceptation\b', _re.IGNORECASE),                    # 9
    _re.compile(r'\beffet\s*de\s*commerce\b', _re.IGNORECASE),          # 10 (nouveau)
    _re.compile(r'\bbillet\s*[àa]\s*ordre\b', _re.IGNORECASE),          # 11 (nouveau)
    _re.compile(r'\bsouscripteur\b', _re.IGNORECASE),                   # 12 (nouveau)

    # Chèque
    _re.compile(r'\bpayez\s*contre\s*ce\s*ch', _re.IGNORECASE),         # 13
    _re.compile(r'\bch[eè]que\s*n[o°]?\b', _re.IGNORECASE),             # 14
    _re.compile(r'\btitulaire\s*du\s*compte\b', _re.IGNORECASE),        # 15
    _re.compile(r'\bnon\s*endossable\b', _re.IGNORECASE),               # 16
    _re.compile(r'\bch[eè]ques?\b', _re.IGNORECASE),                    # 17
    _re.compile(
        r'\bB\.?I\.?A\.?T\b|\bA\.?T\.?B\b|\bU\.?I\.?B\b|\bS\.?T\.?B\b|'
        r'\bB\.?N\.?A\b|\bB\.?H\b|\bU\.?B\.?C\.?I\b|\bB\.?T\.?E\b|\bB\.?T\.?L\b|'
        r'\bA\.?B\.?C\b|\bQ\.?N\.?B\b|\bBanque\s+de\s+Tunisie\b|\bAmen\s*Bank\b|'
        r'\bAttijari\s*Bank\b|\bZitouna\s*Bank\b|\bWifak\b|\bAl\s*Baraka\b|'
        r'\bBanque\s+Centrale\s+de\s+Tunisie\b|\bBCT\b',
        _re.IGNORECASE
    ),                                                                    # 18
    _re.compile(r'\bmontant\s*en\s*toutes\s*lettres\b', _re.IGNORECASE),  # 19 (nouveau)
    _re.compile(r'\bpayable\s*(au|chez)\b', _re.IGNORECASE),              # 20 (nouveau)
    _re.compile(r'\ba\s*vue\b', _re.IGNORECASE),                          # 21 (nouveau) "payable à vue"
    _re.compile(r'\bcarnet\s*de\s*ch[eè]ques\b', _re.IGNORECASE),         # 22 (nouveau)
    _re.compile(r'\bs[ée]rie\s*[:\-]?\s*[A-Z]{1,3}\d{6,}\b', _re.IGNORECASE),  # 23 (nouveau) numéro type chèque
    _re.compile(r'\b\d{2}\s?\d{2,3}\s?\d{2,3}\s?\d{5,8}\s?\d{1,3}\s?\d{1,3}\b'),  # 24 (nouveau) RIB tunisien ~20 chiffres

    # Facture
    _re.compile(r'\bfacture\b', _re.IGNORECASE),                        # 21
    _re.compile(r'\bfacture\s*n[o°]?\s*[:\-]?\s*\d', _re.IGNORECASE),   # 22
    _re.compile(r'\binvoice\b', _re.IGNORECASE),                        # 23
    _re.compile(r'\bnet\s*[àa]\s*payer\b', _re.IGNORECASE),             # 24
    _re.compile(r'\btotal\s*ttc\b', _re.IGNORECASE),                    # 25
    _re.compile(r'\bmontant\s*ttc\b', _re.IGNORECASE),                  # 26
    _re.compile(r'\btva\b', _re.IGNORECASE),                            # 27

    # Bon de livraison / commande
    _re.compile(r'\bbon\s+de\s+livraison\b', _re.IGNORECASE),           # 28
    _re.compile(r'\bbon\s+de\s+commande\b', _re.IGNORECASE),            # 29
    _re.compile(r'\bdelivery\s+note\b', _re.IGNORECASE),                # 30
    _re.compile(r'\bpurchase\s+order\b', _re.IGNORECASE),               # 31
    _re.compile(r'\bbl\s+n[o°]?\b', _re.IGNORECASE),                    # 32
    _re.compile(r'\bbc\s+n[o°]?\b', _re.IGNORECASE),                    # 33

    # Devis / Proforma
    _re.compile(r'\bdevis\s+n[o°]?\b', _re.IGNORECASE),                 # 34
    _re.compile(r'\bproforma\s+invoice\b', _re.IGNORECASE),             # 35

    # Fiche article
    _re.compile(r'\bfiche\s+technique\b', _re.IGNORECASE),              # 36
    _re.compile(r'\bfiche\s+article\b', _re.IGNORECASE),                # 37
]]

_LABELS_REJET_IMMEDIAT_BOM = [
    "tireur (traite)",            # 0
    "domiciliation (traite)",     # 1
    "lettre de change",           # 2
    "valeur reçue (traite)",      # 3
    "à l'ordre de",               # 4
    "échéance (traite)",          # 5
    "tiré (traite)",              # 6
    "aval (traite)",              # 7
    "traite",                     # 8
    "acceptation (traite)",       # 9
    "effet de commerce (traite)", # 10
    "billet à ordre (traite)",    # 11
    "souscripteur (traite)",      # 12
    "payez contre ce chèque",     # 13
    "chèque n°",                  # 14
    "titulaire du compte",        # 15
    "non endossable (chèque)",    # 16
    "chèque",                     # 17
    "banque (BIAT/ATB/UIB)",      # 18
    "montant en toutes lettres",  # 19
    "payable au/chez",            # 20
    "payable à vue (chèque)",     # 21
    "carnet de chèques",          # 22
    "série chèque",               # 23
    "RIB tunisien (20 chiffres)", # 24
    "facture",                    # 25
    "facture n°",                 # 26
    "invoice",                    # 27
    "net à payer",                # 28
    "total ttc",                  # 29
    "montant ttc",                # 30
    "tva",                        # 31
    "bon de livraison",           # 32
    "bon de commande",            # 33
    "delivery note",              # 34
    "purchase order",             # 35
    "BL n°",                      # 36
    "BC n°",                      # 37
    "devis n°",                   # 38
    "proforma invoice",           # 39
    "fiche technique",            # 40
    "fiche article",              # 41
]

_LABELS_REJETES_BOM = {
    "traite": "Traite (Lettre de Change)", "cheque": "Chèque",
    "facture": "Facture", "bon_livraison": "Bon de Livraison",
    "bon_commande": "Bon de Commande", "devis": "Devis / Proforma",
    "fiche_article": "Fiche Article", "inconnu": "Document inconnu",
}

def _categorie_bom_depuis_patterns(patterns_trouves):
    texte = " ".join(patterns_trouves).lower()
    if any(m in texte for m in (
        "traite", "tireur", "tiré", "échéance", "aval",
        "effet de commerce", "billet à ordre", "souscripteur",
    )):
        return "traite"
    if any(m in texte for m in (
        "chèque", "titulaire du compte", "montant en toutes lettres", "payable au/chez",
        "payable à vue", "carnet de chèques", "série chèque", "rib tunisien",
    )):
        return "cheque"
    if "facture" in texte or "invoice" in texte or "ttc" in texte or "tva" in texte:
        return "facture"          # ← priorité déjà correcte, inchangée
    if "livraison" in texte or "delivery" in texte or "bl n" in texte:
        return "bon_livraison"
    if "commande" in texte or "purchase order" in texte or "bc n" in texte:
        return "bon_commande"
    if "devis" in texte or "proforma" in texte:
        return "devis"
    if "fiche" in texte or "code article" in texte or "groupe d'article" in texte:
        return "fiche_article"
    return "inconnu"


def _detecter_rejet_categorie(texte):
    """
    Point d'entrée UNIQUE pour détecter un document non-BOM
    (chèque, traite, facture, BL, BC, devis, fiche article).

    Combine :
      1. La recherche de patterns forts (regex) → priorité.
      2. Le comptage de signaux "fiche article" (≥ 2 mots-clés structurels)
         quand aucun pattern fort n'a matché.

    Retourne la catégorie détectée (str) ou None si rien de concluant.
    Centraliser cette logique permet de la relancer facilement sur un
    texte OCR "renforcé" en cas d'échec de la 1ère passe.
    """
    if not texte:
        return None

    patterns_trouves = [
        _LABELS_REJET_IMMEDIAT_BOM[i]
        for i, p in enumerate(_PATTERNS_REJET_IMMEDIAT_BOM)
        if p.search(texte)
    ]
    if patterns_trouves:
        return _categorie_bom_depuis_patterns(patterns_trouves)

    texte_lower = texte.lower()
    nb_signaux_article = sum(1 for m in _MOTS_ARTICLE_STRUCTURE if m in texte_lower)
    if nb_signaux_article >= 2:
        return "fiche_article"

    return None


def _titre_rejet_bom(cat):
    return "Document refusé — {}".format(_LABELS_REJETES_BOM.get(cat, "Document refusé"))

def _msg_rejet_bom(cat):
    label = _LABELS_REJETES_BOM.get(cat, cat)
    if cat == "inconnu":
        return (
            "Document non reconnu comme Nomenclature (BOM).\n\n"
            "Aucun indicateur trouvé (Nomenclature, Composants, Matière première, Coût total…).\n"
            "Veuillez soumettre un document Nomenclature imprimé."
        )
    return (
        "Type de document détecté : {0}\n\n"
        "Ce module accepte uniquement les documents Nomenclature (BOM).\n"
        "Le document soumis a été identifié comme : « {0} ».\n\n"
        "Pour une {0} → utilisez le module correspondant."
    ).format(label)
# ──────────────────────────────────────────────────────────────────────
# ENDPOINT 1 : Lancement async
# ──────────────────────────────────────────────────────────────────────

@frappe.whitelist()
#api ocr_bom_pipeline.pipeline_bom(POST ) Recevoir fichier -> sauvegarde temp -> frappe.enqueue job async (ocr + extraction BOM)
def pipeline_bom(file_url="", source_doctype="BOM"):
    """
    Lance le pipeline OCR BOM de façon asynchrone.
    Retourne {"success": True, "async": True, "job_id": "<token>"}
    """
    contenu      = None
    nom_original = None
    content_type = ""

    # ── Résolution du fichier ────────────────────────────────────────
    if file_url:
        site_path = frappe.get_site_path()
        if file_url.startswith("/private/"):
            chemin = os.path.join(site_path, "private", "files",
                                  os.path.basename(file_url))
        elif file_url.startswith("/files/"):
            chemin = os.path.join(site_path, "public", "files",
                                  os.path.basename(file_url))
        else:
            chemin = os.path.join(site_path, "public", file_url.lstrip("/"))

        if not os.path.exists(chemin):
            return {"success": False, "erreur": f"Fichier introuvable : {file_url}"}

        nom_original = os.path.basename(chemin)
        with open(chemin, "rb") as f:
            contenu = f.read()
    else:
        files = frappe.request.files
        if not files or "file" not in files:
            return {"success": False,
                    "erreur": "Aucun fichier reçu (form-data, clé: 'file')."}
        file_obj     = files["file"]
        nom_original = file_obj.filename
        content_type = getattr(file_obj, "content_type", "") or ""
        contenu      = file_obj.read()

    if not nom_original:
        nom_original = "nomenclature.pdf"
    nom_original = secure_filename(nom_original)

    # ── Validation taille ────────────────────────────────────────────
    taille_kb = len(contenu) / 1024
    if taille_kb < 2:
        return {
            "success": False,
            "erreur": (
                f"Fichier trop petit ({taille_kb:.1f} KB). "
                "Un document Nomenclature lisible fait généralement ≥ 30 KB."
            )
        }

    nom_fichier, ext = _corriger_extension(nom_original, content_type, contenu)
    extensions_ok    = [".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".xlsx", ".svg"]
    if ext not in extensions_ok:
        return {
            "success": False,
            "erreur": f"Format '{ext}' non supporté. Acceptés : {', '.join(extensions_ok)}"
        }

    # ── Sauvegarde temporaire ────────────────────────────────────────
    dossier_tmp = os.path.join(frappe.get_site_path(), "private", "files")
    os.makedirs(dossier_tmp, exist_ok=True)
    job_token  = str(uuid.uuid4()).replace("-", "")
    chemin_tmp = os.path.join(dossier_tmp,
                              f"ocr_bom_tmp_{job_token}_{nom_fichier}")
    with open(chemin_tmp, "wb") as f:
        f.write(contenu)

    frappe.cache().set_value(
        f"ocr_bom_status_{job_token}", "en_cours", expires_in_sec=3600
    )

    frappe.enqueue(
        "ocr_intelligent.api.ocr_bom_pipeline._executer_pipeline_bom_job",
        queue="long",
        timeout=600,
        chemin_tmp=chemin_tmp,
        nom_fichier=nom_fichier,
        ext=ext,
        source_doctype=source_doctype,
        uploaded_by=frappe.session.user,
        job_token=job_token,
        file_url=file_url,
    )

    return {"success": True, "async": True, "job_id": job_token}


# ──────────────────────────────────────────────────────────────────────
# ENDPOINT 2 : Polling statut
# ──────────────────────────────────────────────────────────────────────

@frappe.whitelist()
#api ocr_bom_pipeline.get_ocr_bom_statut(GET) Polling cache redis (en_cours / termine / erreur)
def get_ocr_bom_statut(job_id):
    """
    Retourne le statut du job OCR BOM.
    Réponses : "en_cours" | "termine" | "erreur" | "inconnu"
    """
    status = frappe.cache().get_value(f"ocr_bom_status_{job_id}")

    if status == "termine":
        result_raw = frappe.cache().get_value(f"ocr_bom_result_{job_id}")
        if result_raw is None:
            return {"status": "termine",
                    "result": {"success": False, "erreur": "Résultat expiré du cache."}}
        try:
            result = json.loads(result_raw) if isinstance(result_raw, str) else result_raw
        except (ValueError, TypeError):
            result = {"success": False, "erreur": str(result_raw)}
        return {"status": "termine", "result": result}

    if status == "erreur":
        erreur_raw = frappe.cache().get_value(f"ocr_bom_erreur_{job_id}")
        try:
            erreur = json.loads(erreur_raw) if isinstance(erreur_raw, str) else erreur_raw
        except (ValueError, TypeError):
            erreur = erreur_raw or "Erreur inconnue."
        return {"status": "erreur", "erreur": erreur}

    if status == "en_cours":
        return {"status": "en_cours"}

    return {"status": "inconnu"}


# ──────────────────────────────────────────────────────────────────────
# ENDPOINT 3 : Attachement fichier au document
# ──────────────────────────────────────────────────────────────────────

@frappe.whitelist()
#api ocr_bom_pipeline.attacher_fichier_a_document(POST) Attache un fichier existant à un document ERPNext
def attacher_fichier_a_document(doctype, docname, file_url):
    """
    Attache un fichier existant à un document ERPNext.
    Utilisé après l'enregistrement d'un BOM pour lier le fichier OCR original.
    """
    try:
        frappe.logger().info(f"[OCR BOM] attacher_fichier_a_document appelé: {doctype}/{docname}, file_url={file_url}")

        if not doctype or not docname or not file_url:
            return {"success": False, "erreur": "Paramètres manquants"}

        # Vérifier que le document existe
        if not frappe.db.exists(doctype, docname):
            return {"success": False, "erreur": f"Document {doctype}/{docname} introuvable"}

        # Chercher le document File correspondant à l'URL
        frappe.logger().info(f"[OCR BOM] Recherche fichier avec URL: {file_url}")
        file_docs = frappe.get_all(
            "File",
            filters={"file_url": file_url},
            fields=["name", "attached_to_doctype", "attached_to_name", "file_name"],
            limit=1
        )

        frappe.logger().info(f"[OCR BOM] Fichiers trouvés: {len(file_docs)}")
        if file_docs:
            frappe.logger().info(f"[OCR BOM] Fichier trouvé: {file_docs[0]}")

        if not file_docs:
            return {"success": False, "erreur": f"Fichier {file_url} introuvable"}

        file_doc = file_docs[0]

        # Si déjà attaché au bon document, ne rien faire
        if file_doc.get("attached_to_doctype") == doctype and file_doc.get("attached_to_name") == docname:
            frappe.logger().info(f"[OCR BOM] Fichier déjà attaché au bon document")
            return {"success": True, "message": "Fichier déjà attaché"}

        # Attacher le fichier au document en utilisant SQL direct (évite les problèmes ORM)
        frappe.logger().info(f"[OCR BOM] Attachement du fichier {file_doc['name']} à {doctype}/{docname}")
        frappe.db.sql("""
            UPDATE `tabFile`
            SET attached_to_doctype = %s, attached_to_name = %s
            WHERE name = %s
        """, (doctype, docname, file_doc["name"]))
        frappe.db.commit()

        frappe.logger().info(f"[OCR BOM] Fichier attaché avec succès via SQL direct")
        return {
            "success": True,
            "message": f"Fichier attaché à {doctype}/{docname}",
            "file_name": file_doc["name"]
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "OCR BOM - Attachement fichier")
        frappe.logger().error(f"[OCR BOM] Erreur attachement: {str(e)}")
        return {"success": False, "erreur": str(e)}


# ──────────────────────────────────────────────────────────────────────
# JOB ASYNCHRONE PRINCIPAL
# ──────────────────────────────────────────────────────────────────────

def _executer_pipeline_bom_job(chemin_tmp, nom_fichier, ext,
                                source_doctype, uploaded_by, job_token, file_url=""):
    """
    Worker principal du pipeline BOM/Nomenclature (exécuté par Redis).

    Rôle : Analyser un fichier de nomenclature et en extraire l'en-tête
    (article produit, société, quantité) et la liste des composants.

    Étapes :
      1. Détection du type de fichier :
         - .xlsx → excel_bom_parser (parser structuré, haute confiance)
           Avantage : conserve la structure tableau, confiance=95.
         - image/PDF → OCR classique + bom_extractor (regex)
      2. Validation minimale : rejet si 0 composant détecté.
      3. Mapping champs_ocr → fieldnames ERPNext BOM via MAPPING_BOM.
      4. Création/mise à jour OCR Document avec tous les champs.
      5. Stockage résultat dans Redis + nettoyage fichier tmp.

    Note Excel (v9.3 fix) : ocr_engine._xlsx() concatène le texte à plat,
    perdant la structure tableau. Le parser dédié excel_bom_parser.py
    utilise openpyxl.iter_rows() pour préserver les colonnes.

    Note détection (v9.4 fix) : la détection de document rejeté
    (chèque/traite/fiche article/facture/BL/...) passe désormais par
    _detecter_rejet_categorie(), rejouée sur le texte OCR renforcé si la
    1ère passe ne conclut à rien (voir bloc IMAGE/PDF ci-dessous).
    """
    try:
        # ══════════════════════════════════════════════════════════════
        # NOUVELLE LOGIQUE : Détection Excel → Parser structuré
        # ══════════════════════════════════════════════════════════════

        if ext == ".xlsx":
            # ── NOUVEAU : construire un texte plat pour détecter le type ──
            try:
                import openpyxl
                wb_check = openpyxl.load_workbook(chemin_tmp, data_only=True, read_only=True)
                cellules = []
                for ws in wb_check.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell is not None:
                                cellules.append(str(cell))
                texte_excel_check = " ".join(cellules)
            except Exception:
                texte_excel_check = ""

            if texte_excel_check:
                # ── si signaux BOM forts présents, on ne rejette JAMAIS ──
                texte_lower = texte_excel_check.lower()
                a_des_signaux_bom_forts = any(s in texte_lower for s in _SIGNAUX_BOM)

                if not a_des_signaux_bom_forts:
                    categorie = _detecter_rejet_categorie(texte_excel_check)
                    if categorie:
                        _stocker_resultat(job_token, {
                            "success":       False,
                            "type_document": categorie,
                            "titre":         _titre_rejet_bom(categorie),
                            "erreur":        _msg_rejet_bom(categorie),
                        })
                        return
            # ── EXCEL : Utiliser le parser structuré ─────────────────
            from ocr_intelligent.ocr.excel_bom_parser import extraire_bom_depuis_excel
            try:
                resultat = extraire_bom_depuis_excel(chemin_tmp)
            except Exception:
                frappe.log_error(frappe.get_traceback(), "OCR BOM - excel_bom_parser crash")
                categorie_secours = _detecter_rejet_categorie(texte_excel_check)
                if categorie_secours:
                    _stocker_resultat(job_token, {
                        "success":       False,
                        "type_document": categorie_secours,
                        "titre":         _titre_rejet_bom(categorie_secours),
                        "erreur":        _msg_rejet_bom(categorie_secours),
                    })
                else:
                    _stocker_resultat(job_token, {
                        "success": False,
                        "erreur": "Le fichier Excel ne correspond pas à la structure attendue d'une Nomenclature (BOM)."
                    })
                return

            type_doc    = resultat["type_document"]
            champs_ocr  = resultat["champs"]
            composants  = resultat["composants"]
            confiances  = resultat["confiances"]
            score       = 95

            # Validation minimale
            if not composants:
                _stocker_resultat(job_token, {
                    "success":       False,
                    "type_document": type_doc,
                    "erreur": (
                        "❌ Aucun composant détecté dans le fichier Excel.\n\n"
                        "Le fichier Excel ne contient pas de tableau de composants (BOM Items).\n"
                        "Vérifiez que votre fichier contient :\n"
                        "  • Une section avec les en-têtes : Code Article, Nom Article, Quantité, UdM\n"
                        "  • Au moins une ligne de composant sous les en-têtes\n\n"
                        "Champs header extraits : " + ", ".join(champs_ocr.keys()) if champs_ocr else "aucun"
                    ),
                })
                return

        else:
            # ── IMAGE/PDF : Utiliser OCR classique ───────────────────
            from ocr_intelligent.ocr.ocr_engine import get_engine
            engine = get_engine()
            try:
                res_ocr = engine.extraire_texte(chemin_tmp)
            except Exception as e:
                frappe.log_error(frappe.get_traceback(), "OCR BOM Pipeline")
                _stocker_erreur(job_token, f"Erreur OCR : {e}")
                return

            texte_brut = res_ocr.get("texte", "")
            score      = res_ocr.get("score_confiance", 0)

            frappe.log_error(
                message=texte_brut[:3000],
                title=f"OCR BOM DEBUG - {nom_fichier}"
            )

            # ── Vérification texte minimal ───────────────────────────
            mots = [m for m in (texte_brut or "").split() if len(m) > 1]
            if len(mots) < 5:
                _stocker_resultat(job_token, {
                    "success":        False,
                    "score_confiance": score,
                    "erreur": (
                        f"Aucun texte détecté (score OCR {score}%). "
                        "Utilisez une image nette ≥ 300 DPI."
                    )
                })
                return

            # ── DÉTECTION UNIFIÉE (v9.4) ──────────────────────────────
            # Si le document contient déjà un signal BOM fort, on ne
            # cherche même pas à le rejeter.
            texte_lower = texte_brut.lower()
            a_des_signaux_bom_forts = any(s in texte_lower for s in _SIGNAUX_BOM_FORTS)

            categorie_detectee = None
            if not a_des_signaux_bom_forts:
                categorie_detectee = _detecter_rejet_categorie(texte_brut)

            # ── Passe 2 (OCR renforcé) ────────────────────────────────
            # Utile dans 2 cas :
            #   a) aucun signal BOM trouvé du tout → peut-être un BOM mal scanné
            #   b) aucune catégorie de rejet trouvée → un chèque/traite/fiche
            #      article mal OCR-isé passait avant à tort en "inconnu"
            texte_p2 = None
            if not any(s in texte_lower for s in _SIGNAUX_BOM) or (
                not a_des_signaux_bom_forts and categorie_detectee is None
            ):
                texte_p2 = _ocr_enhanced(chemin_tmp, ext)

            if texte_p2:
                texte_combine = f"{texte_brut}\n{texte_p2}"
                texte_combine_lower = texte_combine.lower()

                if any(s in texte_combine_lower for s in _SIGNAUX_BOM):
                    # Le texte enrichi révèle un vrai BOM → on l'adopte
                    texte_brut = texte_combine
                    score      = max(score, 50)
                    categorie_detectee = None
                elif categorie_detectee is None and not a_des_signaux_bom_forts:
                    # Retenter la classification de rejet sur le texte enrichi
                    categorie_detectee = _detecter_rejet_categorie(texte_combine)
                    if categorie_detectee:
                        texte_brut = texte_combine  # garder le meilleur texte pour traçabilité

            if categorie_detectee:
                _stocker_resultat(job_token, {
                    "success":       False,
                    "type_document": categorie_detectee,
                    "titre":         _titre_rejet_bom(categorie_detectee),
                    "erreur":        _msg_rejet_bom(categorie_detectee),
                })
                return

            # ── Extraction des champs (OCR texte) ────────────────────
            from ocr_intelligent.ocr.bom_extractor import extraire_champs_bom
            try:
                resultat = extraire_champs_bom(texte_brut)
            except Exception:
                # Garde-fou anti-crash (v9.4 fix) : avant d'abandonner,
                # on retente une classification de rejet — c'est ce qui
                # manquait et causait l'erreur générique sur les fiches
                # article qui faisaient planter l'extracteur.
                frappe.log_error(frappe.get_traceback(), "OCR BOM - bom_extractor crash")
                categorie_secours = _detecter_rejet_categorie(texte_brut)
                if categorie_secours:
                    _stocker_resultat(job_token, {
                        "success":       False,
                        "type_document": categorie_secours,
                        "titre":         _titre_rejet_bom(categorie_secours),
                        "erreur":        _msg_rejet_bom(categorie_secours),
                    })
                else:
                    _stocker_erreur(job_token, "Erreur interne du pipeline OCR BOM (extraction impossible).")
                return

            type_doc    = resultat["type_document"]
            champs_ocr  = resultat["champs"]
            composants  = resultat["composants"]
            confiances  = resultat["confiances"]

            # Rejet si aucun signal BOM (dernier filet de sécurité)
            if type_doc == "inconnu" and not any(
                s in texte_brut.lower() for s in _SIGNAUX_BOM
            ):
                _stocker_resultat(job_token, {
                    "success":       False,
                    "type_document": "inconnu",
                    "titre":         _titre_rejet_bom("inconnu"),
                    "erreur":        _msg_rejet_bom("inconnu"),
                })
                return

        # ── Étape 5 : Validation des champs Link ─────────────────────
        # Vérifier que l'article existe
        if "item" in champs_ocr:
            raw_item = str(champs_ocr["item"]).strip()
            if not frappe.db.exists("Item", raw_item):
                # Recherche par item_code
                items = frappe.get_all(
                    "Item",
                    filters=[["item_code", "like", f"%{raw_item}%"]],
                    fields=["name", "item_code"],
                    limit=5,
                )
                if items:
                    champs_ocr["item"] = items[0]["item_code"]
                    confiances["item"] = max(confiances.get("item", 0.5) - 0.1, 0.4)

        # Vérifier UOM
        if "uom" in champs_ocr:
            if not frappe.db.exists("UOM", champs_ocr["uom"]):
                uoms = frappe.get_all(
                    "UOM",
                    filters=[["name", "like", f"%{champs_ocr['uom']}%"]],
                    fields=["name"],
                    limit=3,
                )
                if uoms:
                    champs_ocr["uom"] = uoms[0]["name"]
                else:
                    # UdM non trouvée → supprimer pour éviter erreur validation
                    del champs_ocr["uom"]
                    confiances.pop("uom", None)

        # Vérifier Routing (nouveau) - Création automatique si inexistant
        if "routing" in champs_ocr and champs_ocr["routing"]:
            routing_name = str(champs_ocr["routing"]).strip()
            if routing_name and not frappe.db.exists("Routing", routing_name):
                try:
                    # Créer automatiquement le routing
                    routing_doc = frappe.get_doc({
                        "doctype": "Routing",
                        "routing_name": routing_name,
                        "operations": []  # Table vide - à remplir manuellement après
                    })
                    routing_doc.insert(ignore_permissions=True)
                    frappe.db.commit()

                    frappe.logger().info(f"[OCR BOM] Routing créé automatiquement: {routing_name}")
                    confiances["routing"] = 0.95  # Haute confiance (extraction Excel)
                except Exception as e:
                    frappe.logger().error(f"[OCR BOM] Erreur création routing {routing_name}: {str(e)}")
                    # En cas d'erreur, supprimer le champ pour éviter blocage
                    del champs_ocr["routing"]
                    confiances.pop("routing", None)

        # ── Étape 6 : Validation composants ──────────────────────────
        composants_valides = []
        for comp in composants:
            code = comp.get("item_code", "").strip()
            if not code:
                continue

            # Vérifier que l'item existe
            exists = frappe.db.exists("Item", code)
            comp["item_exists"] = bool(exists)

            # Valider et normaliser UOM
            if "uom" in comp and comp["uom"]:
                uom_value = str(comp["uom"]).strip()
                if not frappe.db.exists("UOM", uom_value):
                    # Chercher UOM similaire
                    uoms = frappe.get_all(
                        "UOM",
                        filters=[["name", "like", f"%{uom_value}%"]],
                        fields=["name"],
                        limit=3,
                    )
                    if uoms:
                        comp["uom"] = uoms[0]["name"]
                    else:
                        # UOM non trouvée → utiliser "Unité" par défaut
                        comp["uom"] = "Unité"

            # Valider stock_uom
            if "stock_uom" in comp and comp["stock_uom"]:
                stock_uom_value = str(comp["stock_uom"]).strip()
                if not frappe.db.exists("UOM", stock_uom_value):
                    # Utiliser la même UOM que uom si stock_uom invalide
                    comp["stock_uom"] = comp.get("uom", "Unité")

            composants_valides.append(comp)

   # ── Étape 7 : Créer/Mettre à jour le OCR Document ────────────
        # Combiner header + composants dans extracted_field pour traçabilité
        tous_champs = {**champs_ocr}
        if composants_valides:
            tous_champs["composants"] = composants_valides

        ocr_doc_name = None
        for variante in _variantes_nom(nom_fichier):
            existants = frappe.get_list(
                "OCR Document",
                filters={"document_name": variante},
                fields=["name"],
                limit=1,
            )
            if existants:
                ocr_doc_name = existants[0]["name"]
                break

        # Résumé lisible (remplace le texte brut absent pour Excel/OCR)
        # ── Texte extrait : reconstruction lisible du contenu Excel/BOM ──
        def _construire_texte_extrait(champs, composants):
            lignes = []
            lignes.append("=== EN-TÊTE NOMENCLATURE ===")
            for cle, valeur in champs.items():
                lignes.append(f"{cle} : {valeur}")

            if composants:
                lignes.append("")
                lignes.append("=== COMPOSANTS ===")
                for i, comp in enumerate(composants, start=1):
                    code = comp.get("item_code", "")
                    nom  = comp.get("item_name", "")
                    qty  = comp.get("qty", "")
                    uom  = comp.get("uom", "")
                    rate = comp.get("rate", "")
                    lignes.append(
                        f"{i}. {code} — {nom} | Qté : {qty} {uom} | Prix : {rate}"
                    )

            return "\n".join(lignes) if lignes else "Aucune donnée extraite."

        texte_extrait_complet = _construire_texte_extrait(champs_ocr, composants_valides)

        statut = "Validé" if len(champs_ocr) > 0 else "Validation requise"
        if ocr_doc_name:
            frappe.db.set_value("OCR Document", ocr_doc_name, {
                "extracted_text":  texte_extrait_complet,
                "extracted_field": json.dumps(tous_champs, ensure_ascii=False, indent=2),
                "confidence_score": score,
                "status":          statut,
                "uploaded_by":     uploaded_by,
            })
        else:
            ocr_doc_obj = frappe.get_doc({
                "doctype":          "OCR Document",
                "document_name":    nom_fichier,
                "uploaded_by":      uploaded_by,
                "confidence_score": score,
                "extracted_text":   texte_extrait_complet,
                "extracted_field":  json.dumps(tous_champs, ensure_ascii=False, indent=2),
                "status":           statut,
            })
            ocr_doc_obj.insert(ignore_permissions=True)
            ocr_doc_name = ocr_doc_obj.name

        # ── Étape 7b : Attacher le fichier original ───────────────────
        try:
            deja_joint = frappe.db.exists("File", {
                "attached_to_doctype": "OCR Document",
                "attached_to_name":    ocr_doc_name,
                "file_name":           nom_fichier,
            })
            if not deja_joint:
                with open(chemin_tmp, "rb") as _f:
                    contenu_fichier = _f.read()

                file_doc = frappe.get_doc({
                    "doctype":             "File",
                    "file_name":           nom_fichier,
                    "attached_to_doctype": "OCR Document",
                    "attached_to_name":    ocr_doc_name,
                    "attached_to_field":   "file_url",
                    "is_private":          1,
                    "content":             contenu_fichier,
                })
                file_doc.insert(ignore_permissions=True)

                frappe.db.set_value(
                    "OCR Document", ocr_doc_name,
                    "file_url", file_doc.file_url
                )
        except Exception:
            frappe.log_error(frappe.get_traceback(),
                             "OCR BOM — Échec attachement fichier")

        frappe.db.commit()



        # ── Étape 8 : Construction résultat final ─────────────────────
        _stocker_resultat(job_token, {
            "success":         True,
            "type_document":   type_doc,
            "score_confiance": score,
            "champs_remplis":  champs_ocr,
            "composants":      composants_valides,
            "confiances":      confiances,
            "nb_composants":   len(composants_valides),
            "ocr_document_id": ocr_doc_name,  # ID du OCR Document créé
            "file_url":        file_url,
        })

    except Exception:
        frappe.log_error(frappe.get_traceback(), "OCR BOM Pipeline")
        _stocker_erreur(job_token, "Erreur interne du pipeline OCR BOM.")
    finally:
        # Nettoyage fichier temporaire
        try:
            if os.path.exists(chemin_tmp):
                os.remove(chemin_tmp)
        except OSError:
            pass


# ──────────────────────────────────────────────────────────────────────
# HELPERS - Noms de fichiers
# ──────────────────────────────────────────────────────────────────────

def _nettoyer_nom_frappe(nom):
    """Retire les suffixes hash Frappe (ex: _a1b2c3d4) du nom de fichier."""
    base, ext = os.path.splitext(nom)
    base_nettoye = re.sub(r'[_\-]?[a-f0-9]{6,10}$', '', base, flags=re.IGNORECASE)
    if base_nettoye and base_nettoye != base:
        return base_nettoye + ext
    return nom


def _variantes_nom(nom_fichier):
    """Génère des variantes de noms pour rechercher les OCR Documents existants."""
    variantes = [nom_fichier]
    nom_nettoye = _nettoyer_nom_frappe(nom_fichier)
    if nom_nettoye != nom_fichier:
        variantes.append(nom_nettoye)
    base = os.path.splitext(nom_fichier)[0]
    if base not in variantes:
        variantes.append(base)
    base_nettoye = os.path.splitext(nom_nettoye)[0]
    if base_nettoye not in variantes:
        variantes.append(base_nettoye)
    return variantes


# ──────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────

def _stocker_resultat(job_token, result):
    """Persiste le résultat du job BOM dans Redis (statut 'termine', TTL 1h)."""
    frappe.cache().set_value(
        f"ocr_bom_result_{job_token}",
        json.dumps(result, default=str),
        expires_in_sec=3600
    )
    frappe.cache().set_value(
        f"ocr_bom_status_{job_token}", "termine", expires_in_sec=3600
    )


def _stocker_erreur(job_token, erreur):
    """Persiste un message d'erreur dans Redis et positionne le statut BOM à 'erreur'."""
    frappe.cache().set_value(
        f"ocr_bom_erreur_{job_token}", erreur, expires_in_sec=3600
    )
    frappe.cache().set_value(
        f"ocr_bom_status_{job_token}", "erreur", expires_in_sec=3600
    )


def _corriger_extension(nom, content_type, contenu):
    """Détecte/corrige l'extension selon les magic bytes ou content-type."""
    nom_base, ext = os.path.splitext(nom)
    ext = ext.lower()
    if not ext:
        if content_type:
            ct = content_type.lower()
            if "pdf" in ct:
                ext = ".pdf"
            elif "png" in ct:
                ext = ".png"
            elif "jpeg" in ct or "jpg" in ct:
                ext = ".jpg"
            elif "tiff" in ct:
                ext = ".tiff"
            elif "bmp" in ct:
                ext = ".bmp"
            else:
                ext = ".bin"
        elif contenu[:4] == b"%PDF":
            ext = ".pdf"
        elif contenu[:8] in (b"\x89PNG\r\n\x1a\n",):
            ext = ".png"
        elif contenu[:2] in (b"\xff\xd8",):
            ext = ".jpg"
        else:
            ext = ".bin"
    return f"{nom_base}{ext}", ext


def _ocr_enhanced(chemin, ext):
    """Passe OCR améliorée pour documents difficiles."""
    try:
        import cv2
        import numpy as np
        import pytesseract

        if ext == ".pdf":
            try:
                import fitz
                doc = fitz.open(chemin)
                page = doc[0]
                pix = page.get_pixmap(dpi=300)
                img_arr = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                    pix.height, pix.width, pix.n
                )
                if pix.n == 4:
                    img_arr = cv2.cvtColor(img_arr, cv2.COLOR_RGBA2BGR)
                elif pix.n == 1:
                    img_arr = cv2.cvtColor(img_arr, cv2.COLOR_GRAY2BGR)
            except ImportError:
                return None
        else:
            img_arr = cv2.imread(chemin)

        if img_arr is None:
            return None

        gray = cv2.cvtColor(img_arr, cv2.COLOR_BGR2GRAY)
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        texte = pytesseract.image_to_string(
            binary, lang="fra+eng",
            config="--oem 1 --psm 6"
        )
        return texte
    except Exception:
        return None