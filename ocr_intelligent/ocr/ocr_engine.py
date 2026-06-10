# -*- coding: utf-8 -*-
import cv2
import numpy as np
import pytesseract

try:
    from paddleocr import PaddleOCR
    _PADDLE_AVAILABLE = True
except Exception:
    _PADDLE_AVAILABLE = False

_paddle = None


# ─────────────────────────────
# PADDLE LAZY LOAD
# ─────────────────────────────
def get_paddle():
    global _paddle

    if _paddle is not None:
        return _paddle

    if not _PADDLE_AVAILABLE:
        return None

    _paddle = PaddleOCR(
        use_angle_cls=False,
        lang="fr",
        show_log=False
    )
    return _paddle


# ─────────────────────────────
# DETECTION FLOU (FAST)
# ─────────────────────────────
def _est_flou(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return cv2.Laplacian(gray, cv2.CV_64F).var() < 70


# ─────────────────────────────
# PREPROCESSING LIGHT
# ─────────────────────────────
def _preprocess(img, flou=False):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    if flou:
        # léger sharpen seulement
        gaussian = cv2.GaussianBlur(gray, (0, 0), 2)
        gray = cv2.addWeighted(gray, 1.5, gaussian, -0.5, 0)

    return gray


# ─────────────────────────────
# TESSERACT FAST
# ─────────────────────────────
def _tesseract_fast(img):
    return pytesseract.image_to_string(
        img,
        lang="fra+eng",
        config="--oem 1 --psm 6"  # oem 1 = LSTM only (~30% faster than oem 3)
    )


class OCREngine:

    def extraire_texte(self, path):
        if path.endswith(".pdf"):
            return self._pdf(path)
        elif path.endswith(".xlsx"):
            return self._xlsx(path)
        elif path.endswith(".svg"):
            return self._svg(path)

        return self._image(path)

    # ─────────────────────────────
    # IMAGE (OPTIMISÉ)
    # ─────────────────────────────
    def _image(self, path):
        img = cv2.imread(path)

        if img is None:
            return self._vide()

        img = self._resize(img)

        # 🔍 flou calculé UNE seule fois
        flou = _est_flou(img)

        # ⚡ FAST PATH TESSERACT
        gray = _preprocess(img, flou=False)
        texte_fast = _tesseract_fast(gray)

        if len(texte_fast.split()) > 20:  # 20-word threshold: skip Paddle when text is already rich
            return {
                "texte": texte_fast,
                "score_confiance": 75,
                "moteur": "tesseract_fast",
                "image_floue": flou
            }

        # 🔥 IMAGE FLOUE → traitement léger + retry
        if flou:
            gray_flou = _preprocess(img, flou=True)
            texte_flou = _tesseract_fast(gray_flou)

            if len(texte_flou.split()) > len(texte_fast.split()):
                texte_fast = texte_flou

        # 🚀 PADDLE seulement si nécessaire
        paddle = get_paddle()

        if paddle:
            try:
                res = paddle.ocr(img)
                texte_paddle = self._extract_paddle(res)

                if len(texte_paddle.split()) > len(texte_fast.split()):
                    return {
                        "texte": texte_paddle,
                        "score_confiance": 85,
                        "moteur": "paddle",
                        "image_floue": flou
                    }
            except Exception:
                pass

        return {
            "texte": texte_fast,
            "score_confiance": 65,
            "moteur": "tesseract",
            "image_floue": flou
        }

    # ─────────────────────────────
    # PDF (ULTRA OPTIMISÉ)
    # ─────────────────────────────
    def _pdf(self, path):
        from pdf2image import convert_from_path

        # ⚡ 1 seule page + DPI réduit
        pages = convert_from_path(path, dpi=200, first_page=1, last_page=1)

        textes = []

        for p in pages:
            img = cv2.cvtColor(np.array(p), cv2.COLOR_RGB2BGR)
            res = self._image_array(img)
            textes.append(res["texte"])

        return {
            "texte": "\n".join(textes),
            "score_confiance": 80,
            "moteur": "pdf_fast"
        }

    def _image_array(self, img):
        img = self._resize(img)
        gray = _preprocess(img)
        texte = _tesseract_fast(gray)
        return {"texte": texte}

    # ─────────────────────────────
    # XLSX (EXCEL) - Extract images
    # ─────────────────────────────
    def _xlsx(self, path):
        """Extrait les images d'un fichier Excel et effectue l'OCR"""
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage
            import io
            
            wb = load_workbook(path)
            textes = []
            
            # Parcourir toutes les feuilles
            for sheet in wb.worksheets:
                # Extraire les images
                if hasattr(sheet, '_images'):
                    for img in sheet._images:
                        try:
                            # Convertir l'image en array numpy
                            pil_img = PILImage.open(io.BytesIO(img._data()))
                            cv_img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
                            res = self._image_array(cv_img)
                            textes.append(res["texte"])
                        except Exception:
                            continue
                
                # Extraire aussi le texte des cellules
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            textes.append(cell)
            
            texte_final = "\n".join(filter(None, textes))
            return {
                "texte": texte_final,
                "score_confiance": 70,
                "moteur": "xlsx_extractor"
            }
        except Exception as e:
            return self._vide()

    # ─────────────────────────────
    # SVG - Convert to raster image
    # ─────────────────────────────
    def _svg(self, path):
        """Convertit un SVG en image raster et effectue l'OCR"""
        try:
            from cairosvg import svg2png
            from PIL import Image as PILImage
            import io
            
            # Convertir SVG en PNG en mémoire
            png_data = svg2png(url=path, dpi=300)
            
            # Charger en tant qu'image PIL puis convertir en array OpenCV
            pil_img = PILImage.open(io.BytesIO(png_data))
            cv_img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGBA2BGR)
            
            # Utiliser la méthode _image standard
            res = self._image_array(cv_img)
            return {
                "texte": res["texte"],
                "score_confiance": 75,
                "moteur": "svg_raster"
            }
        except Exception as e:
            return self._vide()

    # ─────────────────────────────
    # RESIZE OPTIMISÉ
    # ─────────────────────────────
    def _resize(self, img):
        h, w = img.shape[:2]

        if w > 2400:  # downscale huge scans — INTER_AREA preserves sharpness
            scale = 2400 / w
            return cv2.resize(img, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)

        if w < 800:  # upscale small images — INTER_CUBIC for quality
            scale = 1200 / w
            return cv2.resize(img, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

        return img

    # ─────────────────────────────
    def _extract_paddle(self, result):
        textes = []

        for page in (result or []):
            for line in page:
                textes.append(line[1][0])

        return "\n".join(textes)

    def _vide(self):
        return {"texte": "", "score_confiance": 0}


# ─────────────────────────────
# SINGLETON
# ─────────────────────────────
_ENGINE = None

def get_engine():
    global _ENGINE

    if _ENGINE is None:
        _ENGINE = OCREngine()

    return _ENGINE