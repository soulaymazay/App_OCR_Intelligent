# -*- coding: utf-8 -*-
"""
excel_bom_parser.py - Groupe Bayoudh Metal

Extracteur spécialisé pour les fichiers Excel de Nomenclature (BOM).
Préserve la structure tabulaire au lieu de convertir en texte plat.

USAGE:
    from ocr_intelligent.ocr.excel_bom_parser import extraire_bom_depuis_excel
    resultat = extraire_bom_depuis_excel(chemin_fichier)
"""

import re
from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════════════
# MAPPINGS ET CONSTANTES
# ══════════════════════════════════════════════════════════════════════

# Mots-clés pour détecter les sections dans l'Excel
_SECTION_KEYWORDS = {
    "header": [
        "article", "nom article", "nomenclature", "bom",
        "société", "societe", "company", "groupe",
        "quantité produite", "quantite produite", "quantity",
        "unité de mesure", "unite de mesure", "uom", "udm",
        "devise", "currency", "tnd",
        "prix basé sur", "rm cost as per", "valuation rate",
        "gamme", "route", "routing",
        "transfert matériel", "transfer material",
        "taux de change", "conversion rate", "exchange rate",
        "est actif", "is active", "actif", "est défaut", "is default",
        "taux de perte", "scrap", "scrap percentage",
    ],
    "composants": [
        "composants", "bom items", "bill of material",
        "matières premières", "matieres premieres", "raw materials",
        "code article", "item code", "article", "code",
        "nom article", "item name", "designation",
    ],
    "operations": [
        "opérations", "operations", "gamme de fabrication",
        "poste de travail", "workstation", "durée",
    ],
    "scrap": [
        "matériaux de rebut", "scrap items", "scrap materials",
        "chutes", "rebut",
    ],
    "couts": [
        "résumé des coûts", "resume des couts", "cost summary",
        "coût matière", "cout matiere", "raw material cost",
        "coût exploitation", "cout exploitation", "operating cost",
        "coût total", "cout total", "total cost",
    ],
}

# Mapping des champs header Excel → ERPNext BOM
_FIELD_MAPPING = {
    "article": "item",
    "article (code)": "item",
    "item code": "item",
    "code article": "item",
    
    "nom de l'article": "item_name",
    "nom article": "item_name",
    "item name": "item_name",
    "désignation": "item_name",
    
    "société": "company",
    "societe": "company",
    "company": "company",
    "groupe": "company",
    
    "quantité produite": "quantity",
    "quantite produite": "quantity",
    "quantity": "quantity",
    "qté": "quantity",
    
    "unité de mesure": "uom",
    "unite de mesure": "uom",
    "uom": "uom",
    "udm": "uom",
    "unité": "uom",
    
    "devise": "currency",
    "currency": "currency",
    
    "prix basé sur": "rm_cost_as_per",
    "prix base sur": "rm_cost_as_per",
    "rm cost as per": "rm_cost_as_per",
    
    "gamme": "routing",
    "routing": "routing",
    "route": "routing",
    
    "transfert matériel contre": "transfer_material_against",
    "transfert materiel contre": "transfer_material_against",
    "transfer material against": "transfer_material_against",
    
    "taux de change": "conversion_rate",
    "conversion rate": "conversion_rate",
    "exchange rate": "conversion_rate",
    
    "est actif": "is_active",
    "is active": "is_active",
    "actif": "is_active",
    
    "est défaut": "is_default",
    "est defaut": "is_default",
    "is default": "is_default",
    "par défaut": "is_default",
    
    "taux de perte": "scrap_percentage",
    "scrap percentage": "scrap_percentage",
    "scrap %": "scrap_percentage",
    "%scrap": "scrap_percentage",
}

# Mapping des colonnes de composants Excel → ERPNext BOM Item
# Note : Ordre important - patterns spécifiques avant génériques
_COMPONENT_FIELD_MAPPING = {
    # Code article (colonne B dans BOM-0004-001)
    "code article": "item_code",
    "code de l'article": "item_code",
    "item code": "item_code",
    "article": "item_code",
    "code": "item_code",
    
    # Nom article (colonne C)
    "nom de l'article": "item_name",
    "nom article": "item_name",
    "item name": "item_name",
    "désignation": "item_name",
    "designation": "item_name",
    "nom": "item_name",
    
    # Description
    "description": "description",
    
    # Quantité (colonne D - "Qté")
    "quantité": "qty",
    "quantite": "qty",
    "quantity": "qty",
    "qté": "qty",
    "qte": "qty",
    "qty": "qty",
    
    # UdM (colonne E)
    "unité de mesure": "uom",
    "unite de mesure": "uom",
    "uom": "uom",
    "udm": "uom",
    "unité": "uom",
    "unite": "uom",
    "u.d.m": "uom",
    
    # Qté/U (colonne F)
    "qté/u": "qty_per_unit",
    "qte/u": "qty_per_unit",
    "qty/unit": "qty_per_unit",
    "qtité/unité": "qty_per_unit",
    "quantité par unité": "qty_per_unit",
    "qte/unit": "qty_per_unit",
    
    # UdM Stock
    "uom stock": "stock_uom",
    "stock uom": "stock_uom",
    "udm stock": "stock_uom",
    
    # Facteur conversion
    "facteur conv.": "conversion_factor",
    "facteur conv": "conversion_factor",
    "facteur conversion": "conversion_factor",
    "conversion factor": "conversion_factor",
    "fact. conv": "conversion_factor",
    
    # Prix (colonne G)
    "prix (tnd)": "rate",
    "prix": "rate",
    "rate": "rate",
    "taux (tnd)": "rate",
    "taux": "rate",
    "montant": "rate",
}

# UOM normalization
_UOM_NORMALIZE = {
    "kilogramme": "Kg", "kg": "Kg", "kilo": "Kg",
    "gramme": "g", "gr": "g",
    "mètre": "m", "metre": "m", "ml": "m",
    "m²": "m²", "m2": "m²", "mètre carré": "m²",
    "m³": "m³", "m3": "m³",
    "litre": "L", "liter": "L", "lt": "L", "l": "L",
    "unité": "Unité", "unite": "Unité", "nos": "Unité",
    "pièce": "Unité", "piece": "Unité", "pcs": "Unité", "pce": "Unité",
}


# ══════════════════════════════════════════════════════════════════════
# FONCTION PRINCIPALE
# ══════════════════════════════════════════════════════════════════════

def extraire_bom_depuis_excel(chemin_fichier):
    """
    Extrait les données de Nomenclature (BOM) depuis un fichier Excel.
    
    Retourne un dict :
    {
        "champs": dict,           # Champs header (item, company, quantity, etc.)
        "composants": list[dict], # Liste des composants BOM
        "confiances": dict,       # Confiances par champ (toujours 1.0 pour Excel)
        "type_document": str,     # "nomenclature"
    }
    """
    try:
        wb = load_workbook(chemin_fichier, data_only=True)
        sheet = wb.active
        
        champs = {}
        composants = []
        confiances = {}
        
        # ═══════════════════════════════════════════════════════════════
        # ÉTAPE 1 : Extraire les champs header (recherche par label)
        # ═══════════════════════════════════════════════════════════════
        
        # Parcourir les 30 premières lignes pour trouver les paires label:valeur
        for row_data in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
            # Convertir en liste pour manipulation
            row = list(row_data)
            
            # Chercher des paires dans toutes les colonnes adjacentes possibles
            for i in range(len(row)):
                if row[i] is None:
                    continue
                    
                cell_label = str(row[i]).strip().lower()
                
                # Nettoyer le label (enlever ":", espaces multiples)
                cell_label = cell_label.replace(":", "").strip()
                cell_label = " ".join(cell_label.split())
                
                # Normaliser les accents
                cell_label = cell_label.replace("é", "e").replace("è", "e").replace("ê", "e")
                cell_label = cell_label.replace("à", "a").replace("ô", "o")
                
                # Chercher un mapping
                for label_fr, field_name in _FIELD_MAPPING.items():
                    if field_name in champs:
                        continue  # Déjà trouvé
                    
                    # Match si le label contient la clé ou vice-versa
                    if label_fr in cell_label or cell_label in label_fr:
                        # Chercher la valeur dans les colonnes suivantes (jusqu'à +3 colonnes)
                        for offset in range(1, min(4, len(row) - i)):
                            cell_value = row[i + offset]
                            value = _extraire_valeur_cellule(cell_value, field_name)
                            if value is not None:
                                champs[field_name] = value
                                confiances[field_name] = 1.0
                                break
                        break
        
        # ═══════════════════════════════════════════════════════════════
        # ÉTAPE 2 : Détecter le tableau de composants
        # ═══════════════════════════════════════════════════════════════
        
        # Trouver la ligne d'en-tête du tableau composants
        header_row_idx = None
        header_mapping = {}  # col_index → field_name
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            # Vérifier si cette ligne contient les en-têtes de colonnes
            labels_found = []
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                
                cell_text = str(cell).strip().lower()
                
                # Ignorer les colonnes "#" (numéro de ligne)
                if cell_text == "#" or cell_text == "n°" or cell_text == "no":
                    continue
                
                # Nettoyer le texte (enlever accents, espaces multiples)
                cell_clean = cell_text.replace("é", "e").replace("è", "e").replace("ê", "e")
                cell_clean = cell_clean.replace("à", "a").replace("ù", "u").replace("ô", "o")
                cell_clean = " ".join(cell_clean.split())  # Normaliser espaces
                
                # Chercher un match dans les mappings
                matched = False
                for label, field in _COMPONENT_FIELD_MAPPING.items():
                    # Match exact ou contenu
                    if label == cell_clean or label in cell_clean or cell_clean in label:
                        # Éviter les doublons (prendre le premier match)
                        if not any(f == field for _, f in labels_found):
                            labels_found.append((col_idx, field))
                            matched = True
                            break
                
            # Si on a trouvé au moins 3 colonnes essentielles (code, nom, qté minimum)
            required_fields = {"item_code", "item_name", "qty"}
            found_fields = {field for _, field in labels_found}
            has_required = len(required_fields & found_fields) >= 2  # Au moins 2 sur 3
            
            if len(labels_found) >= 3 and has_required:
                header_row_idx = row_idx
                header_mapping = {col: field for col, field in labels_found}
                break
        
        # ═══════════════════════════════════════════════════════════════
        # ÉTAPE 3 : Extraire les lignes de composants
        # ═══════════════════════════════════════════════════════════════
        
        if header_row_idx:
            # Parcourir les lignes après l'en-tête
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                # Vérifier si c'est une ligne de données (pas une section suivante)
                if _est_ligne_section(row):
                    break  # On a atteint la section suivante (Opérations, Résumé, etc.)
                
                composant = {}
                for col_idx, field_name in header_mapping.items():
                    if col_idx < len(row):
                        value = _extraire_valeur_cellule(row[col_idx], field_name)
                        if value is not None:
                            composant[field_name] = value
                
                # Valider que le composant a au moins un code article
                if composant.get("item_code"):
                    # Normaliser UOM
                    if "uom" in composant:
                        composant["uom"] = _normaliser_uom(str(composant["uom"]))
                    if "stock_uom" in composant:
                        composant["stock_uom"] = _normaliser_uom(str(composant["stock_uom"]))
                    
                    # Valeurs par défaut
                    if "qty" not in composant:
                        composant["qty"] = 1.0
                    if "rate" not in composant:
                        composant["rate"] = 0.0
                    if "conversion_factor" not in composant:
                        composant["conversion_factor"] = 1.0
                    if "qty_per_unit" not in composant:
                        composant["qty_per_unit"] = composant.get("qty", 1.0)
                    
                    composants.append(composant)
        
        # ═══════════════════════════════════════════════════════════════
        # ÉTAPE 4 : Retourner le résultat
        # ═══════════════════════════════════════════════════════════════
        
        return {
            "champs": champs,
            "composants": composants,
            "confiances": confiances,
            "type_document": "nomenclature",
            "nb_composants": len(composants),
        }
    
    except Exception as e:
        import frappe
        frappe.log_error(f"Erreur extraction Excel BOM : {str(e)}", "Excel BOM Parser")
        return {
            "champs": {},
            "composants": [],
            "confiances": {},
            "type_document": "inconnu",
            "erreur": str(e),
        }


# ══════════════════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ══════════════════════════════════════════════════════════════════════

def _extraire_valeur_cellule(cell_value, field_name):
    """Convertit la valeur de cellule selon le type de champ"""
    if cell_value is None or str(cell_value).strip() == "":
        return None
    
    # Champs numériques
    if field_name in ["quantity", "conversion_rate", "scrap_percentage", "qty", 
                      "rate", "conversion_factor", "qty_per_unit"]:
        try:
            # Gérer les formats tunisiens (virgule = séparateur décimal)
            s = str(cell_value).strip()
            
            # Enlever le symbole % pour scrap_percentage
            if field_name == "scrap_percentage":
                s = s.replace("%", "").strip()
            
            # Enlever espaces et convertir virgule en point
            s = s.replace(" ", "").replace(",", ".")
            return float(s)
        except (ValueError, TypeError):
            return None
    
    # Champs booléens
    if field_name in ["is_active", "is_default"]:
        v = str(cell_value).strip().lower()
        if v in ["oui", "yes", "1", "true", "actif", "vrai"]:
            return 1
        return 0
    
    # Champs texte
    return str(cell_value).strip()


def _est_ligne_section(row):
    """Détecte si une ligne marque le début d'une nouvelle section"""
    if not row:
        return False
    
    # Vérifier les 3 premières cellules
    for cell in row[:3]:
        if cell is None:
            continue
        text = str(cell).strip().lower()
        
        # Mots-clés de sections
        section_keywords = [
            "opérations", "operations", "gamme",
            "résumé des coûts", "cost summary", "coût",
            "matériaux de rebut", "scrap items",
            "paramètres", "parameters",
        ]
        
        for keyword in section_keywords:
            if keyword in text:
                return True
    
    return False


def _normaliser_uom(uom_text):
    """Normalise l'unité de mesure"""
    uom = str(uom_text).strip().lower()
    return _UOM_NORMALIZE.get(uom, uom_text.strip())
