# -*- coding: utf-8 -*-
"""
excel_article_parser.py - Groupe Bayoudh Metal
Parser dédié aux fiches Article au format Excel (Sage ERP export).

Structure attendue (sheet "Fiche Article") :
  Colonne A : Label1    Colonne B : Valeur1    Colonne C : Label2    Colonne D : Valeur2

Sheet "Stock par site" :
  Ligne 1 : titre       Ligne 2 : en-têtes     Lignes 3+ : données

Corrections v2 :
  - Lecture directe du layout 4 colonnes (A=label, B=val, C=label, D=val)
  - item_name extrait depuis "Désignation" colonne D (ligne "Article")
  - item_group : accepte les codes courts comme "CONS" (pas de filtre longueur)
  - stock_uom / purchase_uom / sales_uom : UN → "Unité" via _normaliser_udm
  - Coef UA-US/UV-US 1,000000 → filtrés (pas de conversion)
  - Stock par site : lecture complète avec tous les champs
  - Poids = 0 → non extrait
"""

import re
import json


# ──────────────────────────────────────────────────────────────────────
# UDM MAP (identique à article_extractor.py)
# ──────────────────────────────────────────────────────────────────────
_UDM_MAP = {
    "Unité": ["unité", "unite", "unit", "un", "u", "ea", "each",
              "nos", "pce", "pcs", "pièce", "piece", "qté", "qte"],
    "Kg":    ["kg", "kilogramme", "kilo"],
    "g":     ["g", "gramme", "gr"],
    "m":     ["m", "mètre", "metre", "ml", "mètre linéaire"],
    "m2":    ["m²", "m2", "mètre carré"],
    "m3":    ["m³", "m3", "mètre cube"],
    "L":     ["l", "litre", "liter"],
    "Box":   ["box", "boîte", "boite", "bte", "bt"],
    "Pack":  ["pack", "lot", "set"],
    "Pair":  ["paire", "pair", "pr"],
    "Roll":  ["rouleau", "roll", "rl"],
    "Doz":   ["douzaine", "doz", "dozen"],
}


def _normaliser_udm(val: str) -> str:
    if not val:
        return ""
    val_l = val.strip().lower()
    for udm_erpnext, variantes in _UDM_MAP.items():
        if val_l in variantes:
            return udm_erpnext
    return ""


def _str(v) -> str:
    """Convertit une valeur de cellule en chaîne propre."""
    if v is None:
        return ""
    return str(v).strip()


def _float(v) -> float:
    """Convertit en float (gère virgule décimale FR)."""
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace(" ", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _int(v) -> int:
    return int(_float(v))


# ──────────────────────────────────────────────────────────────────────
# PARSER PRINCIPAL
# ──────────────────────────────────────────────────────────────────────

def extraire_article_depuis_excel(chemin: str) -> dict:
    """
    Lit un fichier Excel de fiche article Sage ERP et retourne
    {"champs": {...}, "confiances": {...}} compatible avec article_extractor.py
    """
    try:
        import openpyxl
    except ImportError:
        return {"erreur": "openpyxl non installé. pip install openpyxl"}

    try:
        wb = openpyxl.load_workbook(chemin, data_only=True)
    except Exception as e:
        return {"erreur": f"Impossible d'ouvrir le fichier Excel : {e}"}

    champs = {}
    confiances = {}

    # ── Sheet "Fiche Article" ─────────────────────────────────────────
    sheet_name = None
    for name in wb.sheetnames:
        if "fiche" in name.lower() or "article" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Construire un dictionnaire {label_normalisé: valeur} depuis le layout 4 colonnes
    label_val = {}  # label bas de casse → valeur brute

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not any(c is not None for c in row):
            continue

        cells = list(row)
        # Étendre à 4 colonnes si besoin
        while len(cells) < 4:
            cells.append(None)

        a, b, c, d = cells[0], cells[1], cells[2], cells[3]

        # Paires (A=label, B=valeur) et (C=label, D=valeur)
        if a is not None:
            key_a = _str(a).lower().strip(": ")
            if key_a:
                label_val[key_a] = b

        if c is not None:
            key_c = _str(c).lower().strip(": ")
            if key_c:
                label_val[key_c] = d

    # ── Extraction des champs depuis label_val ────────────────────────

    # item_code — chercher "article" comme label
    for k in ("article", "code article", "code", "référence", "ref"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val and len(val) >= 2:
                champs["item_code"] = val
                confiances["item_code"] = 0.97
                break

    # item_name — chercher "désignation" comme label
    for k in ("désignation", "designation", "désig", "libellé", "libelle", "nom article"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val and len(val) >= 2:
                champs["item_name"] = val
                confiances["item_name"] = 0.97
                break

    # item_group — chercher "catégorie", "groupe", "famille"
    for k in ("catégorie", "categorie", "groupe", "groupe d'article", "item group", "famille"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                # PAS de filtre longueur — "CONS" est un groupe valide dans Sage
                champs["item_group"] = val
                confiances["item_group"] = 0.90
                break

    # statut → disabled + custom_statut_article
    for k in ("statut", "status"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k]).lower()
            champs["disabled"] = "0" if "actif" in val and "in" not in val else "1"
            champs["custom_statut_article"] = _str(label_val[k]).capitalize()
            confiances["disabled"] = 0.92
            confiances["custom_statut_article"] = 0.92
            break

    # Gestion stock → is_stock_item
    for k in ("gestion stock",):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k]).lower()
            champs["is_stock_item"] = "1" if "géré" in val or "gere" in val else "0"
            champs["custom_gestion_stock"] = _str(label_val[k])
            confiances["is_stock_item"] = 0.88
            confiances["custom_gestion_stock"] = 0.88
            break

    # Stock < 0 autorisé
    for k in ("stock < 0 autorisé", "stock < 0 autorise", "stock negatif"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k]).lower()
            champs["custom_stock_negatif"] = "1" if "oui" in val or "yes" in val else "0"
            confiances["custom_stock_negatif"] = 0.88
            break

    # Gestion série
    for k in ("gestion série", "gestion serie"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k]).lower()
            champs["has_serial_no"] = "0" if "pas" in val or "non" in val else "1"
            champs["custom_gestion_serie_mode"] = _str(label_val[k])
            confiances["has_serial_no"] = 0.85
            break

    # Gestion lot
    for k in ("gestion lot",):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k]).lower()
            champs["has_batch_no"] = "0" if "pas" in val or "non" in val else "1"
            champs["custom_gestion_lot_mode"] = _str(label_val[k])
            confiances["has_batch_no"] = 0.85
            break

    # Compteur lot / série
    for k, champ in [("compteur lot", "custom_compteur_lot"),
                     ("compteur série", "custom_compteur_serie"),
                     ("compteur serie", "custom_compteur_serie")]:
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs[champ] = val
                confiances[champ] = 0.85

    # Gestion péremption
    for k in ("gestion de péremption", "gestion péremption", "gestion peremption",
              "gestion de peremption"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs["custom_gestion_peremption"] = val
                confiances["custom_gestion_peremption"] = 0.85
            break

    # Article remplacement
    for k in ("article remplacement", "article de remplacement"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs["custom_article_remplacement"] = val
                confiances["custom_article_remplacement"] = 0.85
            break

    # Mode gestion / standard
    for k in ("standard", "mode gestion"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs["custom_mode_gestion"] = val
                confiances["custom_mode_gestion"] = 0.82
            break

    # ── Unités ───────────────────────────────────────────────────────

    # Unité stock → stock_uom
    for k in ("unité stock", "unite stock", "unité de stock", "unité de mesure stock"):
        if k in label_val and label_val[k] is not None:
            udm = _normaliser_udm(_str(label_val[k]))
            if udm:
                champs["stock_uom"] = udm
                confiances["stock_uom"] = 0.92
            break

    # Unité conditionnement
    for k in ("unité conditionnement", "unite conditionnement"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                udm = _normaliser_udm(val)
                # Ne stocker que si ce n'est pas "Unité" générique
                if udm and udm not in ("Unité", "Unite"):
                    champs["custom_uom_conditionnement"] = udm
                    confiances["custom_uom_conditionnement"] = 0.85
            break

    # Unité poids + poids US
    for k in ("unité poids", "unite poids"):
        if k in label_val and label_val[k] is not None:
            udm = _normaliser_udm(_str(label_val[k]))
            if udm:
                champs["weight_uom"] = udm
                confiances["weight_uom"] = 0.85
            break

    for k in ("poids de l'us", "poids us", "poids"):
        if k in label_val and label_val[k] is not None:
            v = _float(label_val[k])
            if v and v != 0:
                champs["weight_per_unit"] = str(round(v, 6))
                confiances["weight_per_unit"] = 0.82
            break

    # Unité achat + coef UA-US
    for k in ("unité achat", "unite achat"):
        if k in label_val and label_val[k] is not None:
            udm = _normaliser_udm(_str(label_val[k]))
            if udm:
                champs["purchase_uom"] = udm
                confiances["purchase_uom"] = 0.88
            break

    for k in ("coefficient ua-us", "coef ua-us", "coef. ua-us", "coefficient ua us"):
        if k in label_val and label_val[k] is not None:
            v = _float(label_val[k])
            if v and v != 1.0 and v != 0:
                champs["custom_coef_ua_us"] = str(round(v, 6))
                confiances["custom_coef_ua_us"] = 0.88
            break

    # Unité vente + coef UV-US
    for k in ("unité vente", "unite vente"):
        if k in label_val and label_val[k] is not None:
            udm = _normaliser_udm(_str(label_val[k]))
            if udm:
                champs["sales_uom"] = udm
                confiances["sales_uom"] = 0.88
            break

    for k in ("coefficient uv-us", "coef uv-us", "coef. uv-us", "coefficient uv us"):
        if k in label_val and label_val[k] is not None:
            v = _float(label_val[k])
            if v and v != 1.0 and v != 0:
                champs["custom_coef_uv_us"] = str(round(v, 6))
                confiances["custom_coef_uv_us"] = 0.88
            break

    # Unité statistique + coef Ustat-US
    for k in ("unité statistique", "unite statistique"):
        if k in label_val and label_val[k] is not None:
            udm = _normaliser_udm(_str(label_val[k]))
            # Ne stocker que si différent de "Unité" (valeur par défaut Sage)
            if udm and udm not in ("Unité",):
                champs["custom_uom_stat"] = udm
                confiances["custom_uom_stat"] = 0.85
            break

    for k in ("coefficient ustat-us", "coef ustat-us", "coef. ustat-us"):
        if k in label_val and label_val[k] is not None:
            v = _float(label_val[k])
            if v and v != 1.0 and v != 0:
                champs["custom_coef_ustat_us"] = str(round(v, 6))
                confiances["custom_coef_ustat_us"] = 0.88
            break

    # ── Champs prix ───────────────────────────────────────────────────
    for k in ("prix de vente", "prix vente", "prix de vente standard",
              "prix de base", "standard rate"):
        if k in label_val and label_val[k] is not None:
            v = _float(label_val[k])
            if v and v > 0:
                champs["standard_rate"] = str(round(v, 3))
                confiances["standard_rate"] = 0.92
            break

    for k in ("prix d'achat", "prix achat", "prix d'achat standard",
              "last purchase rate"):
        if k in label_val and label_val[k] is not None:
            v = _float(label_val[k])
            if v and v > 0:
                champs["last_purchase_rate"] = str(round(v, 3))
                confiances["last_purchase_rate"] = 0.92
            break
# ── Familles statistiques — Nature ───────────────────────────────
    for k in ("nature", "famille statistique 1", "fam. stat. 1"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs["custom_nature"] = val
                confiances["custom_nature"] = 0.88
            break

    # ── Onglet Gestion — Famille coût ────────────────────────────────
    for k in ("famille coût", "famille cout", "coût famille",
              "cout famille", "famille de coût", "famille de cout"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs["custom_famille_cout"] = val
                confiances["custom_famille_cout"] = 0.88
            break

    # ── Onglet Comptabilité ──────────────────────────────────────────
    for k in ("code comptable", "compte", "code compta"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs["custom_code_comptable"] = val
                confiances["custom_code_comptable"] = 0.88
            break

    for k in ("niveau taxe", "niveau de taxe", "taxe", "tva"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k])
            if val:
                champs["custom_niveau_taxe"] = val
                confiances["custom_niveau_taxe"] = 0.88
            break

    # ── Onglet Vente — Type composition ──────────────────────────────
    for k in ("type article", "type d'article", "type composition",
              "type de composition"):
        if k in label_val and label_val[k] is not None:
            val = _str(label_val[k]).lower()
            if "nomenclature" in val:
                champs["custom_type_composition"] = "Composé nomenclature"
            elif "kit" in val:
                champs["custom_type_composition"] = "Composé kit"
            else:
                champs["custom_type_composition"] = "Article normal"
            confiances["custom_type_composition"] = 0.85
            break

    # ── Unités achat et vente ─────────────────────────────────────────
    for k in ("unité achat", "unite achat", "ua", "unité d'achat"):
        if k in label_val and label_val[k] is not None:
            udm = _normaliser_udm(_str(label_val[k]))
            if udm:
                champs["purchase_uom"] = udm
                confiances["purchase_uom"] = 0.88
            break

    for k in ("unité vente", "unite vente", "uv", "unité de vente"):
        if k in label_val and label_val[k] is not None:
            udm = _normaliser_udm(_str(label_val[k]))
            if udm:
                champs["sales_uom"] = udm
                confiances["sales_uom"] = 0.88
            break
    # ── Sheet "Stock par site" ────────────────────────────────────────
    site_sheet = None
    for name in wb.sheetnames:
        if "site" in name.lower() or "stock" in name.lower():
            site_sheet = name
            break

    if site_sheet and site_sheet != sheet_name:
        ws_sites = wb[site_sheet]
        rows = list(ws_sites.iter_rows(min_row=1, values_only=True))

        # Chercher la ligne d'en-têtes (celle qui contient "Site" ou "Code")
        header_row_idx = None
        headers = []
        for i, row in enumerate(rows):
            vals = [_str(c).lower() for c in row if c is not None]
            if "site" in vals or "code" in vals:
                header_row_idx = i
                headers = [_str(c).lower() for c in row]
                break

        if header_row_idx is not None:
            # Index des colonnes
            def col_idx(names):
                for name in names:
                    for i, h in enumerate(headers):
                        if name in h:
                            return i
                return -1

            idx_code    = col_idx(["site", "code"])
            idx_label   = col_idx(["nom site", "nom", "label", "désignation"])
            idx_abc     = col_idx(["abc", "catégorie abc"])
            idx_inv     = col_idx(["inventaire", "mode inventaire"])
            idx_retrait = col_idx(["retrait", "mode retrait"])
            idx_achet   = col_idx(["acheteur", "approv"])
            idx_secu    = col_idx(["sécurité", "securite", "sécu"])
            idx_max     = col_idx(["maximum", "max"])
            idx_seuil   = col_idx(["seuil", "réappro", "reappro"])
            idx_qmin    = col_idx(["qté mini", "qte mini", "quantité mini", "quantite mini"])

            sites_appro = []
            for row in rows[header_row_idx + 1:]:
                if not any(c is not None for c in row):
                    continue
                cells = list(row)

                def cell(idx):
                    if idx < 0 or idx >= len(cells):
                        return None
                    return cells[idx]

                code = _str(cell(idx_code))
                if not code:
                    continue

                label   = _str(cell(idx_label)) if idx_label >= 0 else ""
                abc     = _str(cell(idx_abc)) if idx_abc >= 0 else ""
                inv     = _str(cell(idx_inv)) if idx_inv >= 0 else ""
                retrait = _str(cell(idx_retrait)) if idx_retrait >= 0 else ""
                achet   = _str(cell(idx_achet)) if idx_achet >= 0 else ""
                secu    = _int(cell(idx_secu)) if idx_secu >= 0 else 0
                maxi    = _int(cell(idx_max)) if idx_max >= 0 else 0
                seuil   = _int(cell(idx_seuil)) if idx_seuil >= 0 else 0
                qmin    = _int(cell(idx_qmin)) if idx_qmin >= 0 else 0

                sites_appro.append({
                    "code":                code,
                    "label":               label,
                    "safety_stock":        secu,
                    "stock_max":           maxi,
                    "reorder_level":       seuil,
                    "reorder_qty":         qmin,
                    "lot_size":            0,
                    "categorie_abc":       abc,
                    "mode_inventaire":     inv,
                    "mode_retrait":        retrait,
                    "gestion_emplacement": False,
                    "acheteur_approv":     achet,
                })

            if sites_appro:
                champs["custom_sites_appro"] = json.dumps(sites_appro, ensure_ascii=False)
                confiances["custom_sites_appro"] = 0.92

                # Remonter les valeurs non-nulles dans les champs flat
                non_nul = lambda field: [s[field] for s in sites_appro if s.get(field)]
                if non_nul("safety_stock"):
                    champs["safety_stock"] = str(max(non_nul("safety_stock")))
                    confiances["safety_stock"] = 0.88
                if non_nul("stock_max"):
                    champs["custom_stock_max"] = str(max(non_nul("stock_max")))
                    confiances["custom_stock_max"] = 0.88
                if non_nul("reorder_level"):
                    champs["custom_seuil_reappro"] = str(max(non_nul("reorder_level")))
                    confiances["custom_seuil_reappro"] = 0.88
                if non_nul("reorder_qty"):
                    champs["custom_qte_mini_reappro"] = str(max(non_nul("reorder_qty")))
                    confiances["custom_qte_mini_reappro"] = 0.88

                # Catégorie ABC et mode inventaire depuis la 1ère ligne non-vide
                for site in sites_appro:
                    if site.get("categorie_abc") and "custom_categorie_abc" not in champs:
                        champs["custom_categorie_abc"] = site["categorie_abc"]
                        confiances["custom_categorie_abc"] = 0.88
                    if site.get("mode_inventaire") and "custom_mode_inventaire" not in champs:
                        champs["custom_mode_inventaire"] = site["mode_inventaire"]
                        confiances["custom_mode_inventaire"] = 0.88
                    if site.get("mode_retrait") and "custom_mode_retrait" not in champs:
                        champs["custom_mode_retrait"] = site["mode_retrait"]
                        confiances["custom_mode_retrait"] = 0.85
                    if site.get("acheteur_approv") and "custom_acheteur_approv" not in champs:
                        champs["custom_acheteur_approv"] = site["acheteur_approv"]
                        confiances["custom_acheteur_approv"] = 0.80

    return {
        "champs":     champs,
        "confiances": confiances,
    }


# ──────────────────────────────────────────────────────────────────────
# TEST LOCAL
# ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys, json
    chemin = sys.argv[1] if len(sys.argv) > 1 else "Fiche_Article_ACP0011.xlsx"
    res = extraire_article_depuis_excel(chemin)
    print(json.dumps(res, ensure_ascii=False, indent=2))